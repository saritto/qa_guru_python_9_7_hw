import zipfile
import csv
from pypdf import PdfReader
from io import BytesIO, TextIOWrapper
from openpyxl import load_workbook


# проверка pdf файла
def test_read_pdf():
    with zipfile.ZipFile('resources/file.zip', 'r') as file_zip:
        with file_zip.open('PDF.pdf') as text_pdf:
            pdf_data = BytesIO(text_pdf.read())
            pdf_file = PdfReader(pdf_data)
            number_of_pages = len(pdf_file.pages)
            page = pdf_file.pages[3]
            pdf_text = page.extract_text()
            assert number_of_pages == 84, "Количество страниц в PDF-file не 84!"
            assert "БР = браузерные расширения" in pdf_text


# проверка csv файла
def test_read_csv():
    with (zipfile.ZipFile('resources/file.zip', 'r') as file_zip):
        with file_zip.open('Sample.csv', 'r') as csv_file:
            csv_reader = list(csv.reader(TextIOWrapper(csv_file, encoding="latin-1")))
            assert "Eldon Base for stackable storage shelf, platinum" in csv_reader[0][1], "Текст не соответствует!"
            assert len(csv_reader) == 10, "Количество строк в CSV-file не 10!"


# проверка xlsx файла
def test_read_xlsx():
    with (zipfile.ZipFile('resources/file.zip', 'r') as file_zip):
        with file_zip.open('sample3.xlsx') as xls_file:
            workbook = load_workbook(filename=xls_file)
            sheet = workbook.active
            sheet_names = workbook.sheetnames
            assert "January" in sheet['B2'].value, "Указанный текст не соответствует тексту в XLSX-file!"
            assert "MyLinks" in sheet_names[1], "Указанное название не соответствует названию листа в XLSX-file!"
