import zipfile
import shutil
import csv
from pypdf import PdfReader
from io import BytesIO, TextIOWrapper
from openpyxl import load_workbook


# создание архива
files_zip = zipfile.ZipFile("tmp/file.zip", 'w')
files_zip.write("tmp/Sample.csv", arcname='Sample.csv', compress_type=zipfile.ZIP_DEFLATED)
files_zip.write("tmp/sample3.xlsx", arcname='sample3.xlsx', compress_type=zipfile.ZIP_DEFLATED)
files_zip.write("tmp/Браузерные_расширения.pdf", arcname='Браузерные_расширения.pdf', compress_type=zipfile.ZIP_DEFLATED)


# перемещение в resources
shutil.move("tmp/file.zip", "resources/file.zip")


# чтение из архива
text_csv = files_zip.read('Sample.csv')
text_pdf = files_zip.read('Браузерные_расширения.pdf')
text_xlsx = files_zip.read('sample3.xlsx')


# проверка pdf файла
pdf_file = PdfReader(BytesIO(text_pdf))
number_of_pages = len(pdf_file.pages)
page = pdf_file.pages[3]
pdf_text = page.extract_text()
assert number_of_pages == 84, "Количество страниц в PDF-file не 84!"
assert "БР = браузерные расширения" in pdf_text, "Указанный текст не соответствует тексту в PDF-file!"
files_zip.close()


# проверка csv файла
with zipfile.ZipFile('resources/file.zip', 'r') as file_zip:
    with file_zip.open('Sample.csv', 'r') as csv_file:
        csv_reader = list(csv.reader(TextIOWrapper(csv_file, encoding="latin-1")))
        assert "Eldon Base for stackable storage shelf, platinum" in csv_reader[0][1], "Указанный текст не соответствует тексту в CSV-file!"
        assert len(csv_reader) == 10, "Количество строк в CSV-file не 10!"


# проверка xlsx файла
    with file_zip.open('sample3.xlsx') as xls_file:
        workbook = load_workbook(filename=xls_file)
        sheet = workbook.active
        sheet_names = workbook.sheetnames
        assert "January" in sheet['B2'].value, "Указанный текст не соответствует тексту в XLSX-file!"
        assert "MyLinks" in sheet_names[1], "Указанное название не соответствует названию листа в XLSX-file!"

